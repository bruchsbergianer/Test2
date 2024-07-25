# -*- coding: utf-8 -*-
"""
Created on Tue Mar 19 11:33:24 2024

@author: harderd
"""

import openpyxl
import tkinter

from tkinter import filedialog
from datetime import datetime
from tkinter import ttk

# Dateiauswahldialog öffnen
root = tkinter.Tk()
root.withdraw()
file_path = filedialog.askopenfilename(title="Excel-Datei auswählen", filetypes=[("Excel-Dateien", "*.xlsx")])

# Excel-Datei laden
wb = openpyxl.load_workbook(file_path)

# Arbeitsblatt "Tabelle1" auswählen
ws = wb["Tabelle1"]

# Summe der Minuten in Spalte D berechnen
summe_minuten = 0
for row in ws.iter_rows(min_row=2):  # erste Zeile überspringen
    zelle_d = row[3].value
    if zelle_d is not None:
        if zelle_d is not None:
            start_zeit = datetime.strptime(zelle_d[:5], "%H:%M")
            end_zeit = datetime.strptime(zelle_d[8:], "%H:%M")
            zeitspanne = end_zeit - start_zeit
            minuten = zeitspanne.total_seconds() / 60
            summe_minuten += minuten


# Anzahl der eindeutigen Namen in Spalte C zählen
namen = set()
for row in ws.iter_rows(min_row=2):  # erste Zeile überspringen
    zelle_c = row[6].value
    if zelle_c is not None:
        namen.add(zelle_c)

# Ergebnisse ausgeben
#print(f"Summe der Minuten: {summe_minuten} h")
#print(f"Anzahl der eindeutigen Namen: {len(namen)}")

root = tkinter.Tk()
frm = ttk.Frame(root, padding=10)
frm.grid()
ttk.Label(frm, text=f"Summe der Stunden: {summe_minuten/60} h").grid(column=0, row=0)
ttk.Label(frm, text=f"Anzahl der eindeutigen Namen: {len(namen)}").grid(column=0, row=1)
ttk.Button(frm, text="Quit", command=root.destroy).grid(column=1, row=2)
root.mainloop()